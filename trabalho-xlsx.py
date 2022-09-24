from openpyxl import load_workbook
import numpy  as np
import time

start_time = time.time()
wb = load_workbook(filename='FECHAMENTO_MAIS__NEGOCIADAS_5minutos.xlsx', data_only=True)
ws = wb['Plan1']

allCells = np.array([[cell.value for cell in row] for row in ws.iter_rows()])

firstCol = 2
firstRow = 2
nCols = 12
nRows = 12
max = ws.max_row-12

while firstRow <= max:
    try:
        m = allCells[(firstRow-1):(firstRow+nRows-1), 1:13].astype('float64')
        if np.isfinite(np.sum(m)):
            invm = np.linalg.inv(m)
            print(invm, end="\n\n")
        else:
            print("Matriz possui dados NAN!\n\n")
    except np.linalg.LinAlgError as err:
        print(str(err) + "\n\n")
    except ValueError as err:
        print("Não tem como trabalhar com uma matriz que possui dados que não são números\n\n")
    except:
        print("Erro!\n\n")
    firstRow = firstRow+1

print("--- %s seconds ---" % (time.time() - start_time))